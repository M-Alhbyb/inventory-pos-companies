"""Transaction management views with import/export functionality."""

from django.shortcuts import render, redirect
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from tablib import Dataset
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from base.models import Transaction
from base.resources import TransactionResource

ITEMS_PER_PAGE = 10

@login_required
def transactions_view(request):
    """Display all transactions with filtering, search, and pagination."""
    # Redirect representatives to their own transactions page
    if request.user.user_type == 'representative':
        return redirect('base:rep_transactions')
    
    transactions = Transaction.objects.select_related('user').prefetch_related('items__product').order_by('-date')
    
    # Handle search
    q = request.GET.get('q')
    if q:
        transactions = transactions.filter(
            Q(id__icontains=q) |
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )
    
    # Handle transaction type filter
    transaction_type = request.GET.get('type')
    if transaction_type and transaction_type != 'all':
        transactions = transactions.filter(type=transaction_type)
    
    # Handle user type filter
    user_type = request.GET.get('user_type')
    if user_type and user_type != 'all':
        transactions = transactions.filter(user__user_type=user_type)
    
    # Get statistics
    stats = {
        'total_transactions': transactions.count(),
        'total_take': transactions.filter(type='take').count(),
        'total_payment': transactions.filter(type='payment').count(),
        'total_restore': transactions.filter(type='restore').count(),
        'total_fees': transactions.filter(type='fees').count(),
        'total_pending': transactions.filter(status='pending').count(),
    }
    
    # Handle status filter
    status = request.GET.get('status')
    if status and status != 'all':
        transactions = transactions.filter(status=status)

    # Handle POST requests (export/import)
    if request.method == 'POST':
        if 'export' in request.POST:
            return export_transactions(transactions)
        elif 'import' in request.POST and request.FILES.get('file'):
            return import_transactions(request)

    # Pagination
    paginator = Paginator(transactions, ITEMS_PER_PAGE)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Handle AJAX requests for infinite scroll
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return get_more_data_json(page_obj)

    context = {
        'page_obj': page_obj,
        **stats,
        'current_type': transaction_type or 'all',
        'current_user_type': user_type or 'all',
        'current_status': status or 'all',
        'search_query': q or '',
    }
    
    return render(request, 'transactions.html', context)


def get_more_data_json(page_obj):
    """Return paginated transaction data as JSON for infinite scroll."""
    data = []
    for transaction in page_obj:
        # Get items as list of dicts
        items = []
        for item in transaction.items.all():
            items.append({
                'product_name': item.product.name if item.product else '-',
                'quantity': item.quantity,
                'price': str(item.price) if item.price else '-',
                'total': str(item.total) if item.total else '-',
            })
        
        data.append({
            'id': transaction.id,
            'user_id': transaction.user.id if transaction.user else None,
            'user': transaction.user.username if transaction.user else None,
            'user_type': transaction.user.user_type if transaction.user else None,
            'user_full_name': transaction.user.get_full_name() if transaction.user else None,
            'type': transaction.type,
            'type_display': transaction.get_type_display(),
            'date': transaction.date.strftime('%Y/%m/%d') if transaction.date else None,
            'time': transaction.date.strftime('%I:%M %p') if transaction.date else None,
            'items': items,
            'items_count': len(items),
            'amount': str(transaction.amount) if transaction.amount else None,
        })
    
    return JsonResponse({
        'success': True,
        'data': data,
        'has_next': page_obj.has_next(),
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
    })


def export_transactions(queryset):
    """Export transactions to Excel with styling."""
    resource = TransactionResource()
    dataset = resource.export(queryset)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    
    # Create workbook with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "المعاملات"
    ws.sheet_view.rightToLeft = True  # RTL for Arabic
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    row_fill_light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Column widths
    column_widths = {'A': 15, 'B': 20, 'C': 18, 'D': 22, 'E': 40, 'F': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Write headers
    headers = dataset.headers or ['رقم المعاملة', 'المستخدم', 'نوع المعاملة', 'التاريخ', 'العناصر', 'المجموع']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # Write data rows
    for row_num, row in enumerate(dataset, start=2):
        row_fill = row_fill_light if row_num % 2 == 0 else row_fill_white
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = center_alignment if col_num in [1, 3, 6] else cell_alignment
    
    wb.save(response)
    return response


def import_transactions(request):
    """Import transactions from uploaded file."""
    resource = TransactionResource()
    dataset = Dataset()
    
    uploaded_file = request.FILES['file']
    file_format = uploaded_file.name.split('.')[-1].lower()
    
    try:
        if file_format == 'csv':
            dataset.load(uploaded_file.read().decode('utf-8'), format='csv')
        elif file_format in ['xlsx', 'xls']:
            dataset.load(uploaded_file.read(), format='xlsx')
        else:
            messages.error(request, 'صيغة الملف غير مدعومة. يرجى استخدام ملف CSV أو Excel.')
            return redirect('base:transactions')
        
        # Dry run first to check for errors
        result = resource.import_data(dataset, dry_run=True)
        
        if result.has_errors():
            error_messages = []
            for row_num, errors in result.row_errors():
                for error in errors:
                    error_messages.append(f"الصف {row_num}: {error.error}")
            messages.error(request, f'حدثت أخطاء أثناء الاستيراد: {"; ".join(error_messages[:5])}')
        else:
            # Actually import the data
            result = resource.import_data(dataset, dry_run=False)
            messages.success(request, f'تم استيراد {result.totals["new"]} معاملة جديدة بنجاح!')
            
    except Exception as e:
        messages.error(request, f'فشل الاستيراد: {str(e)}')
    
    return redirect('base:transactions')


@login_required
def approve_transaction(request, transaction_id):
    """Approve a pending transaction (accountant only)."""
    if request.user.user_type != 'accountant':
        messages.error(request, 'غير مصرح لك بهذا الإجراء.')
        return redirect('base:transactions')
    
    try:
        transaction = Transaction.objects.get(pk=transaction_id)
        if transaction.approve():
            messages.success(request, f'تمت الموافقة على المعاملة #{transaction_id} بنجاح.')
        else:
            messages.warning(request, 'هذه المعاملة ليست معلقة.')
    except Transaction.DoesNotExist:
        messages.error(request, 'المعاملة غير موجودة.')
    
    return redirect('base:transactions')


@login_required
def reject_transaction(request, transaction_id):
    """Reject a pending transaction (accountant only)."""
    if request.user.user_type != 'accountant':
        messages.error(request, 'غير مصرح لك بهذا الإجراء.')
        return redirect('base:transactions')
    
    try:
        transaction = Transaction.objects.get(pk=transaction_id)
        if transaction.reject():
            messages.success(request, f'تم رفض المعاملة #{transaction_id}.')
        else:
            messages.warning(request, 'هذه المعاملة ليست معلقة.')
    except Transaction.DoesNotExist:
        messages.error(request, 'المعاملة غير موجودة.')
    
    return redirect('base:transactions')
